"""Processamento leve dos dados e gravação em .xlsx (aba detalhe + resumos)."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def prepare_detail_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    """Padroniza texto e datas para leitura no Excel."""
    out = df.copy()
    for col in out.columns:
        if pd.api.types.is_datetime64_any_dtype(out[col]):
            out[col] = pd.to_datetime(out[col], errors="coerce")
        elif out[col].dtype == object:
            out[col] = out[col].apply(
                lambda x: x.strip() if isinstance(x, str) else x
            )
    cols_renamed = {
        c: _friendly_column_name(c) for c in out.columns
    }
    out = out.rename(columns=cols_renamed)
    return out


def _friendly_column_name(name: str) -> str:
    mapping = {
        "data_vistoria": "Data vistoria",
        "hora_vistoria": "Hora vistoria",
        "tipo_operacao": "Tipo operação",
        "placa": "Placa",
        "tag": "TAG",
        "tecnico_executante_1": "Técnico 1",
        "tecnico_executante_2": "Técnico 2",
        "tecnico_executante_3": "Técnico 3",
        "estado": "UF",
        "cliente": "Cliente",
        "url_download": "URL download",
        "veiculo": "Veículo",
        "tipo_contrato": "Tipo contrato",
        "tipo_equipamento": "Tipo equipamento",
    }
    return mapping.get(str(name).strip(), str(name).replace("_", " ").title())


def _find_column(df: pd.DataFrame, *needles: str) -> str | None:
    lower = {c.lower(): c for c in df.columns}
    for n in needles:
        if n.lower() in lower:
            return lower[n.lower()]
    return None


def _value_counts_clean(
    series: pd.Series, axis_name: str, count_col: str
) -> pd.DataFrame:
    """Contagem por valor, ignorando vazio/NA."""
    vc = (
        series.dropna()
        .astype(str)
        .str.strip()
        .replace("", pd.NA)
        .dropna()
        .value_counts()
        .rename_axis(axis_name)
        .reset_index(name=count_col)
        .sort_values(count_col, ascending=False)
    )
    return vc


def _executor_columns_ordered(df: pd.DataFrame) -> list[tuple[str, str]]:
    """Retorna (rótulo, nome_coluna) para tecnico_executante_1..3 na ordem."""
    labels = (
        ("Executante 1", "tecnico_executante_1"),
        ("Executante 2", "tecnico_executante_2"),
        ("Executante 3", "tecnico_executante_3"),
    )
    out: list[tuple[str, str]] = []
    for rotulo, needle in labels:
        col = _find_column(df, needle)
        if col:
            out.append((rotulo, col))
    return out


def build_summary_frames(
    df: pd.DataFrame, report_title: str
) -> list[tuple[str, pd.DataFrame]]:
    """Monta blocos (título + tabela) para a aba Resumo."""
    sections: list[tuple[str, pd.DataFrame]] = []

    meta = pd.DataFrame(
        [
            {"Campo": "Relatório", "Valor": report_title},
            {
                "Campo": "Gerado em",
                "Valor": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
            },
            {"Campo": "Total de registros", "Valor": len(df)},
        ]
    )
    sections.append(("Visão geral", meta))

    col_cliente = _find_column(df, "cliente")
    if col_cliente:
        por_cliente = _value_counts_clean(df[col_cliente], "Cliente", "Quantidade")
        sections.append(("Por cliente", por_cliente))

    col_estado = _find_column(df, "estado")
    if col_estado:
        por_uf = _value_counts_clean(df[col_estado], "UF", "Quantidade")
        sections.append(("Por UF", por_uf))

    for rotulo_exec, col_tec in _executor_columns_ordered(df):
        sub = df[col_tec]
        por_col = _value_counts_clean(sub, "Técnico", "Quantidade")
        if not por_col.empty:
            sections.append((f"Por técnico ({rotulo_exec})", por_col))

    cols_tec_todas = [
        c
        for c in df.columns
        if "tecnico" in c.lower() or "técnico" in c.lower()
    ]
    if cols_tec_todas:
        partes = []
        for c in cols_tec_todas:
            s = df[c].dropna().astype(str).str.strip()
            s = s[s != ""]
            partes.append(s)
        if partes:
            todas = pd.concat(partes, ignore_index=True)
            por_tec = (
                todas.value_counts()
                .rename_axis("Técnico")
                .reset_index(name="Participações")
                .sort_values("Participações", ascending=False)
            )
            sections.append(
                ("Por técnico (consolidado — todas as colunas)", por_tec)
            )

    col_tipo_op = _find_column(df, "tipo_operacao")
    if col_tipo_op:
        por_tipo = _value_counts_clean(
            df[col_tipo_op], "Tipo operação", "Quantidade"
        )
        sections.append(("Por tipo de operação", por_tipo))

    return sections


def _autosize_columns(ws, max_width: int = 55) -> None:
    for idx, col in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
        length = 0
        for cell in col:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        letter = get_column_letter(idx)
        ws.column_dimensions[letter].width = min(max_width, max(10, length + 2))


def write_report_excel(
    df_raw: pd.DataFrame,
    output_path: Path,
    report_title: str,
) -> Path:
    """
    Grava Excel com abas Detalhe e Resumo.
    Retorna o caminho absoluto do arquivo.
    """
    output_path = output_path.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    detail = prepare_detail_dataframe(df_raw)
    sections = build_summary_frames(df_raw, report_title)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        detail.to_excel(writer, sheet_name="Detalhe", index=False)

        start = 0
        for section_title, sdf in sections:
            pd.DataFrame([[section_title]]).to_excel(
                writer,
                sheet_name="Resumo",
                header=False,
                index=False,
                startrow=start,
            )
            cell = writer.sheets["Resumo"].cell(row=start + 1, column=1)
            cell.font = Font(bold=True, size=12)
            start += 1
            sdf.to_excel(
                writer,
                sheet_name="Resumo",
                startrow=start,
                index=False,
            )
            start += len(sdf) + 2

    _format_workbook(output_path)
    return output_path


def _format_workbook(path: Path) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path)
    for name in wb.sheetnames:
        ws = wb[name]
        _autosize_columns(ws)
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = cell.value
                    cell.font = Font(color="0563C1", underline="single")
                    cell.alignment = Alignment(wrap_text=True, vertical="top")
    wb.save(path)
